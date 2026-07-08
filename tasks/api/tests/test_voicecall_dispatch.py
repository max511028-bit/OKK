"""Тесты для доработок обзвона: шаблон Excel, загрузка/ручной ввод с
known_answers + предпроверкой, диспетчер очереди (poll/claim/result),
воронка и экспорт отчёта.

Сценарий tander-sterlitamak-pack берётся из voicecall/scenarios/*.json
(файловый fallback _vt_load_scenario) — в тестовой БД voicecall_scripts
пуста, так что это ровно тот путь который реально используется.
"""
import io
import json

SCENARIO = "tander-sterlitamak-pack"


def _xlsx_bytes(headers, rows):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class TestUploadTemplate:
    def test_returns_xlsx_with_crit_columns(self, client):
        r = client.get("/voicecall/upload-template", params={"scenario_id": SCENARIO})
        assert r.status_code == 200
        assert "spreadsheet" in r.headers["content-type"]
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        assert headers[0] == "Имя"
        assert headers[1] == "Телефон"
        assert len(headers) > 2  # хотя бы один вопрос сценария

    def test_required_column_is_highlighted(self, client):
        """Телефон — единственное жёстко обязательное поле, его заголовок
        должен быть визуально выделен цветом в скачиваемом шаблоне."""
        r = client.get("/voicecall/upload-template", params={"scenario_id": SCENARIO})
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        name_cell, phone_cell = ws["A1"], ws["B1"]
        assert phone_cell.fill.start_color.rgb not in (None, "00000000")
        assert name_cell.fill.start_color.rgb in (None, "00000000")


class TestPreviewContactsFile:
    def test_preview_returns_all_rows_without_committing(self, client):
        content = _xlsx_bytes(["Имя", "Телефон"], [["Иван", "79991110011"], ["Пётр", "79991110022"]])
        r = client.post(
            "/voicecall/preview-contacts-file",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"scenario_id": SCENARIO, "source": "manual"},
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert len(body["rows"]) == 2
        assert body["rows"][0]["name"] == "Иван"
        assert body["rows"][0]["phone"] == "79991110011"
        # Предпросмотр ничего не должен коммитить в БД
        camps = client.get("/voicecall/campaigns").json()["items"]
        assert len(camps) == 0

    def test_preview_reports_duplicate_phone(self, client):
        """Раньше повторный телефон в файле молча схлопывался без единого
        слова об этом — выглядело как «загрузка потеряла часть строк»."""
        content = _xlsx_bytes(["Имя", "Телефон"], [["А", "79991110033"], ["Б", "79991110033"]])
        r = client.post(
            "/voicecall/preview-contacts-file",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"scenario_id": SCENARIO, "source": "manual"},
        )
        body = r.json()
        assert len(body["rows"]) == 1
        assert body["skipped_duplicate_phone"] == 1

    def test_preview_then_manual_entry_full_flow(self, client):
        """Итоговый сценарий использования: предпросмотр → правка на
        портале → отправка через manual-entry (не через upload-contacts)."""
        content = _xlsx_bytes(["Имя", "Телефон"], [["Иван", "79991110044"]])
        r = client.post(
            "/voicecall/preview-contacts-file",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"scenario_id": SCENARIO, "source": "manual"},
        )
        rows = r.json()["rows"]
        rows[0]["name"] = "Иван Исправленный"  # оператор поправил в таблице

        r2 = client.post("/voicecall/manual-entry", json={
            "campaign_name": "Из предпросмотра",
            "scenario_id": SCENARIO,
            "source": "manual",
            "rows": rows,
        })
        assert r2.status_code == 200, r2.text
        cid = r2.json()["campaign_id"]
        contacts = client.get("/voicecall/contacts", params={"campaign_id": cid}).json()["items"]
        assert contacts[0]["name"] == "Иван Исправленный"


class TestUploadContactsPrecheck:
    def test_stop_factor_screens_out_without_calling(self, client):
        """Пол=женский в файле — известный стоп-фактор, контакт не должен
        попасть в очередь на звонок."""
        content = _xlsx_bytes(["Имя", "Телефон", "Пол"], [["Мария", "79991112233", "женский"]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("test.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Тест отсева", "scenario_id": SCENARIO, "source": "manual"},
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["screened_out"] == 1
        assert body["queued"] == 0

        contacts = client.get("/voicecall/contacts",
                               params={"campaign_id": body["campaign_id"]}).json()["items"]
        assert len(contacts) == 1
        assert contacts[0]["status"] == "skipped"
        assert contacts[0]["screen_out_reason"] is not None

    def test_fully_known_no_stop_marked_done_without_call(self, client):
        """Все ответы известны, стопа нет — контакт сразу 'done' без
        реального звонка (подтверждённое решение пользователя)."""
        headers = ["Имя", "Телефон", "intro", "Пол", "Возраст 18-45",
                   "Гражданство РФ", "Судимости 158/228/105", "День + ночь",
                   "Физнагрузка", "ЛМК", "Когда выйти", "Опыт"]
        row = ["Иван", "79992223344", "да", "мужской", "27", "да", "нет",
               "день+ночь", "да", "да", "завтра", "да"]
        content = _xlsx_bytes(headers, [row])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("test.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Тест полного знания", "scenario_id": SCENARIO, "source": "manual"},
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["precheck_done"] == 1
        assert body["queued"] == 0

        contacts = client.get("/voicecall/contacts",
                               params={"campaign_id": body["campaign_id"]}).json()["items"]
        assert contacts[0]["status"] == "done"
        assert contacts[0]["verdict"] == "passed"
        assert contacts[0]["validation_id"] is not None

    def test_partial_or_unknown_stays_pending(self, client):
        content = _xlsx_bytes(["Имя", "Телефон"], [["Пётр", "79993334455"]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("test.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Тест очереди", "scenario_id": SCENARIO, "source": "manual"},
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["queued"] == 1
        contacts = client.get("/voicecall/contacts",
                               params={"campaign_id": body["campaign_id"]}).json()["items"]
        assert contacts[0]["status"] == "pending"


class TestManualEntry:
    def test_manual_entry_same_precheck_logic(self, client):
        r = client.post("/voicecall/manual-entry", json={
            "campaign_name": "Ручной тест",
            "scenario_id": SCENARIO,
            "source": "manual",
            "rows": [
                {"name": "Ольга", "phone": "79994445566", "Пол": "женский"},
                {"name": "Сергей", "phone": "79995556677"},
            ],
        })
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["total"] == 2
        assert body["screened_out"] == 1
        assert body["queued"] == 1


class TestDispatchQueue:
    def _upload_one_pending(self, client, phone="79996667788"):
        content = _xlsx_bytes(["Имя", "Телефон"], [["Кандидат", phone]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("test.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Диспетчер тест", "scenario_id": SCENARIO, "source": "manual"},
        )
        return r.json()["campaign_id"]

    def test_start_dispatch_requires_password(self, client):
        cid = self._upload_one_pending(client)
        r = client.post(f"/voicecall/campaigns/{cid}/start-dispatch")
        assert r.status_code == 403

    def test_full_dispatch_flow(self, client, portal_token):
        cid = self._upload_one_pending(client, phone="79997778899")
        auth = {"X-Auth-Token": portal_token}

        # Нет заявки на обзвон -> poll ничего не находит
        r = client.get("/voicecall/dispatch/poll", headers=auth)
        assert r.json()["campaign_id"] is None

        # Жмём "Начать обзвон"
        r = client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        assert r.status_code == 200, r.text
        assert r.json()["pending"] == 1

        # Повторный старт пока не done/idle -> не 409, т.к. ещё 'requested' не 'running'... но пул сам его переключит
        # Агент опрашивает и забирает кампанию
        r = client.get("/voicecall/dispatch/poll", headers=auth)
        assert r.status_code == 200
        poll = r.json()
        assert poll["campaign_id"] == cid
        assert poll["scenario_id"] == SCENARIO

        # Пока кампания running - повторный start-dispatch должен быть отклонён
        r = client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        assert r.status_code == 409

        # Агент забирает контакт
        r = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid}, headers=auth)
        assert r.status_code == 200
        claim = r.json()
        assert claim["contact_id"] is not None
        assert claim["phone"] == "79997778899"

        contacts = client.get("/voicecall/contacts", params={"campaign_id": cid}).json()["items"]
        assert contacts[0]["status"] == "calling"
        assert contacts[0]["attempts"] == 1

        # Очередь пуста на второй claim -> кампания переходит в done
        r = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid}, headers=auth)
        assert r.json()["contact_id"] is None
        camps = client.get("/voicecall/campaigns").json()["items"]
        camp = next(c for c in camps if c["id"] == cid)
        assert camp["dispatch_state"] == "done"

        # Агент шлёт результат звонка
        r = client.post("/voicecall/dispatch/result", headers=auth, json={
            "contact_id": claim["contact_id"],
            "status": "answered_completed",
            "verdict": "passed",
            "stop_reason": None,
            "answers": {"Пол": "мужской"},
            "notes": {},
            "transcript": [{"who": "bot", "text": "Привет", "ts": "2026-01-01T00:00:00"}],
            "duration_s": 42.0,
        })
        assert r.status_code == 200, r.text

        contacts = client.get("/voicecall/contacts", params={"campaign_id": cid}).json()["items"]
        assert contacts[0]["status"] == "done"
        assert contacts[0]["last_call_status"] == "answered_completed"
        assert contacts[0]["verdict"] == "passed"
        assert contacts[0]["validation_id"] is not None

    def test_result_voicemail_maps_to_failed(self, client, portal_token):
        cid = self._upload_one_pending(client, phone="79998889900")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        claim = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                             headers=auth).json()

        r = client.post("/voicecall/dispatch/result", headers=auth, json={
            "contact_id": claim["contact_id"],
            "status": "voicemail",
            "error": "автоответчик поймали",
        })
        assert r.status_code == 200
        contacts = client.get("/voicecall/contacts", params={"campaign_id": cid}).json()["items"]
        assert contacts[0]["status"] == "failed"
        assert contacts[0]["last_call_status"] == "voicemail"

    def test_live_transcript_visible_during_call_then_cleared(self, client, portal_token):
        """Живой мониторинг: пока звонок идёт, агент шлёт снимки транскрипта,
        портал их отдаёт по /live; после результата звонка запись удаляется
        (иначе следующий звонок этому же контакту покажет старьё)."""
        cid = self._upload_one_pending(client, phone="79999990011")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        claim = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                             headers=auth).json()
        contact_id = claim["contact_id"]

        assert client.get(f"/voicecall/contacts/{contact_id}/live").json()["transcript"] == []

        r = client.post("/voicecall/dispatch/live", headers=auth, json={
            "contact_id": contact_id,
            "transcript": [{"who": "bot", "text": "Привет", "ts": "2026-01-01T00:00:00"}],
        })
        assert r.status_code == 200, r.text
        live = client.get(f"/voicecall/contacts/{contact_id}/live").json()
        assert live["transcript"] == [{"who": "bot", "text": "Привет", "ts": "2026-01-01T00:00:00"}]

        client.post("/voicecall/dispatch/result", headers=auth, json={
            "contact_id": contact_id, "status": "answered_completed", "verdict": "passed",
        })
        live_after = client.get(f"/voicecall/contacts/{contact_id}/live").json()
        assert live_after["transcript"] == []

    def test_live_endpoint_requires_password(self, client):
        r = client.post("/voicecall/dispatch/live", json={"contact_id": 1, "transcript": []})
        assert r.status_code == 403


class TestHangupMidCall:
    def _claim_one(self, client, portal_token, phone):
        content = _xlsx_bytes(["Имя", "Телефон"], [["Кандидат", phone]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Обрыв тест", "scenario_id": SCENARIO, "source": "manual"},
        )
        cid = r.json()["campaign_id"]
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        claim = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                             headers=auth).json()
        return cid, claim["contact_id"], auth

    def test_partial_answers_preserved_and_dropped_step_recorded(self, client, portal_token):
        cid, contact_id, auth = self._claim_one(client, portal_token, "79991230001")

        r = client.post("/voicecall/dispatch/result", headers=auth, json={
            "contact_id": contact_id,
            "status": "hangup_by_candidate",
            "verdict": None,
            "answers": {"intro": "да"},
            "notes": {},
            "transcript": [{"who": "bot", "text": "Привет", "ts": "2026-01-01T00:00:00"},
                            {"who": "candidate", "text": "да", "ts": "2026-01-01T00:00:01"}],
            "dropped_at_step": "Пол",
        })
        assert r.status_code == 200, r.text

        contacts = client.get("/voicecall/contacts", params={"campaign_id": cid}).json()["items"]
        assert contacts[0]["status"] == "done"
        assert contacts[0]["verdict"] is None
        assert contacts[0]["validation_id"] is not None

        funnel = client.get(f"/voicecall/campaigns/{cid}/funnel").json()
        assert funnel["dropped_mid_call"] == 1
        assert funnel["reached_end"] == 0

        detail = client.get(f"/voicecall/contacts/{contact_id}/detail").json()
        assert detail["dropped_at_step"] == "Пол"
        assert detail["live_answers"] == {"intro": "да"}
        assert len(detail["transcript"]) == 2


class TestCallHistoryAndRecording:
    """По одному контакту может быть несколько попыток дозвона (недозвон,
    потом достучались) — каждая должна остаться в истории с собственным
    транскриптом, а не только последняя (что было раньше)."""

    def _campaign_with_one_contact(self, client, phone="79993330001"):
        content = _xlsx_bytes(["Имя", "Телефон"], [["Кандидат", phone]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "История звонков тест", "scenario_id": SCENARIO, "source": "manual"},
        )
        return r.json()["campaign_id"]

    def _claim_and_result(self, client, auth, campaign_id, status, **extra):
        claim = client.post("/voicecall/dispatch/claim", params={"campaign_id": campaign_id},
                             headers=auth).json()
        contact_id = claim["contact_id"]
        body = {"contact_id": contact_id, "status": status}
        body.update(extra)
        r = client.post("/voicecall/dispatch/result", headers=auth, json=body)
        assert r.status_code == 200, r.text
        return contact_id

    def test_no_answer_then_answered_both_kept_in_history(self, client, portal_token):
        cid = self._campaign_with_one_contact(client)
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)

        # Попытка 1: недозвон — раньше вообще не оставляла следа в истории
        contact_id = self._claim_and_result(client, auth, cid, "no_answer")

        # Контакт возвращается в очередь для второй попытки
        client.post(f"/voicecall/contacts/{contact_id}/retry")
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)

        # Попытка 2: дозвонились и поговорили
        claim2 = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                              headers=auth).json()
        assert claim2["contact_id"] == contact_id
        client.post("/voicecall/dispatch/result", headers=auth, json={
            "contact_id": contact_id, "status": "answered_completed", "verdict": "passed",
            "answers": {"intro": "да"},
            "transcript": [{"who": "bot", "text": "Привет", "ts": "2026-01-01T00:00:00"}],
            "call_session_id": 555111,
        })

        detail = client.get(f"/voicecall/contacts/{contact_id}/detail").json()
        assert len(detail["history"]) == 2
        assert detail["history"][0]["call_status"] == "no_answer"
        assert detail["history"][1]["call_status"] == "answered_completed"
        assert detail["history"][1]["answers"] == {"intro": "да"}

    def test_recording_attached_to_latest_attempt(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79993330002")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(
            client, auth, cid, "answered_completed",
            verdict="passed", call_session_id=999222)

        r = client.post("/voicecall/dispatch/recording", headers=auth, json={
            "contact_id": contact_id,
            "recording_url": "https://app.novofon.ru/system/media/talk/999222/abc/",
        })
        assert r.status_code == 200, r.text

        detail = client.get(f"/voicecall/contacts/{contact_id}/detail").json()
        assert detail["history"][-1]["recording_url"] == "https://app.novofon.ru/system/media/talk/999222/abc/"

    def test_recording_endpoint_requires_password(self, client):
        r = client.post("/voicecall/dispatch/recording",
                         json={"contact_id": 1, "recording_url": "https://x"})
        assert r.status_code == 403

    def test_recheck_transcript_attached_to_latest_attempt(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79993330003")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(
            client, auth, cid, "answered_completed", verdict="passed")

        r = client.post("/voicecall/dispatch/recheck-transcript", headers=auth, json={
            "contact_id": contact_id,
            "recheck_transcript": "[Дорожка 1] да мужской двадцать девять\n\n[Дорожка 2] здравствуйте меня зовут",
        })
        assert r.status_code == 200, r.text

        detail = client.get(f"/voicecall/contacts/{contact_id}/detail").json()
        assert "Дорожка 1" in detail["history"][-1]["recheck_transcript"]

    def test_recheck_transcript_requires_password(self, client):
        r = client.post("/voicecall/dispatch/recheck-transcript",
                         json={"contact_id": 1, "recheck_transcript": "x"})
        assert r.status_code == 403

    def test_corrected_answers_merged_into_latest_attempt(self, client, portal_token):
        """Пере-валидация критичных ответов по записи (2026-07-08):
        уточнённые значения (возраст/стоп-факторы, потерянные маленькой
        моделью на тихой линии) мержатся в answers последней попытки, live
        сохраняется в тексте — не молчаливая перезапись."""
        cid = self._campaign_with_one_contact(client, phone="79993330011")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(
            client, auth, cid, "answered_completed", verdict="passed",
            answers={"Возраст 18-45": 20, "Судимости": "не распознано: не", "Пол": "да"})

        r = client.post("/voicecall/dispatch/recheck-transcript", headers=auth, json={
            "contact_id": contact_id,
            "recheck_transcript": "[Дорожка 1] двадцать девять не было",
            "needs_review": True,
            "review_note": "Возраст: реалтайм «20», по записи «29» — проверьте.",
            "corrected_answers": {
                "Возраст 18-45": "29 (по записи; в реальном времени распознано: 20)",
                "Судимости": "нет (восстановлено по записи)",
            },
        })
        assert r.status_code == 200, r.text
        assert set(r.json()["corrected"]) == {"Возраст 18-45", "Судимости"}

        detail = client.get(f"/voicecall/contacts/{contact_id}/detail").json()
        h = detail["history"][-1]
        assert h["answers"]["Возраст 18-45"] == "29 (по записи; в реальном времени распознано: 20)"
        assert h["answers"]["Судимости"] == "нет (восстановлено по записи)"
        assert h["answers"]["Пол"] == "да"  # не тронуто
        assert h["needs_review"] is True

    def test_recheck_without_corrected_answers_leaves_answers_intact(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79993330012")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(
            client, auth, cid, "answered_completed", verdict="passed",
            answers={"Возраст 18-45": 33})

        client.post("/voicecall/dispatch/recheck-transcript", headers=auth, json={
            "contact_id": contact_id,
            "recheck_transcript": "[Дорожка 1] тридцать три",
        })
        detail = client.get(f"/voicecall/contacts/{contact_id}/detail").json()
        assert detail["history"][-1]["answers"]["Возраст 18-45"] == 33

    def test_needs_review_flag_stored_and_visible_in_detail(self, client, portal_token):
        """Пункт 7 доработок 2026-07: агент может пометить попытку на
        ручную проверку, если пакетная перепроверка записи не
        подтверждает live-вердикт стопа."""
        cid = self._campaign_with_one_contact(client, phone="79993330004")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(
            client, auth, cid, "answered_completed", verdict="stopped")

        r = client.post("/voicecall/dispatch/recheck-transcript", headers=auth, json={
            "contact_id": contact_id,
            "recheck_transcript": "[Дорожка 1] нет не было судимостей",
            "needs_review": True,
            "review_note": "Автосверка: причина «Судимости: да» не подтверждается записью.",
        })
        assert r.status_code == 200, r.text

        detail = client.get(f"/voicecall/contacts/{contact_id}/detail").json()
        last = detail["history"][-1]
        assert last["needs_review"] is True
        assert "не подтверждается" in last["review_note"]

    def test_needs_review_defaults_to_false(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79993330005")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(
            client, auth, cid, "answered_completed", verdict="passed")

        client.post("/voicecall/dispatch/recheck-transcript", headers=auth, json={
            "contact_id": contact_id,
            "recheck_transcript": "[Дорожка 1] да мужской",
        })
        detail = client.get(f"/voicecall/contacts/{contact_id}/detail").json()
        assert detail["history"][-1]["needs_review"] is False
        assert detail["history"][-1]["review_note"] is None

    def test_needs_review_visible_in_contacts_list(self, client, portal_token):
        """Бейдж должен быть виден в основной таблице контактов, не
        только в модалке карточки — иначе рекрутёр его пропустит."""
        cid = self._campaign_with_one_contact(client, phone="79993330006")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(
            client, auth, cid, "answered_completed", verdict="stopped")
        client.post("/voicecall/dispatch/recheck-transcript", headers=auth, json={
            "contact_id": contact_id,
            "recheck_transcript": "[Дорожка 1] нет не было",
            "needs_review": True,
            "review_note": "Проверить вручную.",
        })

        items = client.get(f"/voicecall/contacts?campaign_id={cid}").json()["items"]
        found = next(c for c in items if c["id"] == contact_id)
        assert found["needs_review"] in (True, 1)

    def test_callback_requested_becomes_failed_status_no_auto_requeue(self, client, portal_token):
        """Живой кандидат попросил перезвонить (2026-07): статус ведёт
        себя ровно как no_answer/busy — уходит в failed, last_call_status
        сохраняет точную причину. НИКАКОГО авто-перезвона: контакт не
        возвращается в pending сам по себе, повторный набор только
        вручную кнопкой «Заново» (это гарантируется тем, что мы не
        трогаем сам статус contact дальше failed)."""
        cid = self._campaign_with_one_contact(client, phone="79993330007")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(client, auth, cid, "callback_requested")

        items = client.get(f"/voicecall/contacts?campaign_id={cid}").json()["items"]
        found = next(c for c in items if c["id"] == contact_id)
        assert found["status"] == "failed"
        assert found["last_call_status"] == "callback_requested"

        # Контакт НЕ появляется повторно в claim() сам по себе — очередь
        # pending пуста, кампания просто завершена, никто не перезвонил.
        claim_again = client.post("/voicecall/dispatch/claim",
                                   params={"campaign_id": cid}, headers=auth).json()
        assert claim_again["contact_id"] is None

    def test_callback_requested_counted_in_funnel(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79993330008")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        self._claim_and_result(client, auth, cid, "callback_requested")

        funnel = client.get(f"/voicecall/campaigns/{cid}/funnel").json()
        assert funnel["callback_requested"] == 1


class TestSuspectVoicemails:
    """Пункт 3 доработок 2026-07: конвейер пополнения фраз-автоответчиков —
    GET /voicecall/campaigns/{cid}/suspect-voicemails собирает попытки,
    где словарная детекция скорее всего пропустила заглушку."""

    def _campaign_with_one_contact(self, client, phone="79990000000"):
        content = _xlsx_bytes(["Имя", "Телефон"], [["Тест", phone]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Заглушки тест", "scenario_id": SCENARIO, "source": "manual"},
        )
        return r.json()["campaign_id"]

    def _claim_and_result(self, client, auth, campaign_id, status, **extra):
        claim = client.post("/voicecall/dispatch/claim", params={"campaign_id": campaign_id},
                             headers=auth).json()
        contact_id = claim["contact_id"]
        body = {"contact_id": contact_id, "status": status}
        body.update(extra)
        r = client.post("/voicecall/dispatch/result", headers=auth, json=body)
        assert r.status_code == 200, r.text
        return contact_id

    def test_low_recognition_call_shows_up_with_recheck_transcript(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79994440001")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        contact_id = self._claim_and_result(client, auth, cid, "low_recognition",
                                             call_session_id=1001)
        client.post("/voicecall/dispatch/recheck-transcript", headers=auth, json={
            "contact_id": contact_id,
            "recheck_transcript": "[Дорожка 1] аппарат абонента временно недоступен попробуйте позже",
        })

        r = client.get(f"/voicecall/campaigns/{cid}/suspect-voicemails", headers=auth)
        assert r.status_code == 200, r.text
        items = r.json()["items"]
        assert len(items) == 1
        assert items[0]["contact_id"] == contact_id
        assert "аппарат абонента" in items[0]["recheck_transcript"]

    def test_answered_completed_all_unrecognized_shows_up(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79994440002")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        self._claim_and_result(
            client, auth, cid, "answered_completed", verdict="passed",
            answers={"Шаг 1": "не распознано", "Пол": "не распознано: шум"})

        items = client.get(f"/voicecall/campaigns/{cid}/suspect-voicemails", headers=auth).json()["items"]
        assert len(items) == 1
        assert items[0]["all_answers_unrecognized"] is True

    def test_answered_completed_with_real_answers_not_shown(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79994440003")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        self._claim_and_result(
            client, auth, cid, "answered_completed", verdict="passed",
            answers={"Шаг 1": "да", "Пол": "не распознано"})

        items = client.get(f"/voicecall/campaigns/{cid}/suspect-voicemails", headers=auth).json()["items"]
        assert items == []

    def test_requires_password(self, client, portal_token):
        cid = self._campaign_with_one_contact(client, phone="79994440004")
        r = client.get(f"/voicecall/campaigns/{cid}/suspect-voicemails")
        assert r.status_code == 403

    def test_unknown_campaign_404(self, client, portal_token):
        auth = {"X-Auth-Token": portal_token}
        r = client.get("/voicecall/campaigns/999999/suspect-voicemails", headers=auth)
        assert r.status_code == 404


class TestPauseResume:
    def _campaign_with_two_pending(self, client, phone1, phone2):
        content = _xlsx_bytes(["Имя", "Телефон"], [["A", phone1], ["B", phone2]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Пауза тест", "scenario_id": SCENARIO, "source": "manual"},
        )
        return r.json()["campaign_id"]

    def test_pause_blocks_next_claim_current_call_unaffected(self, client, portal_token):
        cid = self._campaign_with_two_pending(client, "79997770001", "79997770002")
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)
        claim1 = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                              headers=auth).json()
        assert claim1["contact_id"] is not None

        r = client.post(f"/voicecall/campaigns/{cid}/pause-dispatch", headers=auth)
        assert r.status_code == 200

        # Второй контакт не выдаётся, пока на паузе
        claim2 = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                              headers=auth).json()
        assert claim2["contact_id"] is None
        assert claim2.get("paused") is True

        # Пауза — не "конец обзвона"
        camps = client.get("/voicecall/campaigns").json()["items"]
        camp = next(c for c in camps if c["id"] == cid)
        assert camp["dispatch_state"] == "running"
        assert camp["dispatch_paused"] is True

        # poll() тоже не должен пытаться повторно подхватить кампанию на паузе
        assert client.get("/voicecall/dispatch/poll", headers=auth).json()["campaign_id"] is None

        # Первый звонок как ни в чём не бывало доигрывается результатом
        r = client.post("/voicecall/dispatch/result", headers=auth, json={
            "contact_id": claim1["contact_id"], "status": "answered_completed", "verdict": "passed",
        })
        assert r.status_code == 200

        r = client.post(f"/voicecall/campaigns/{cid}/resume-dispatch", headers=auth)
        assert r.status_code == 200

        claim2b = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                               headers=auth).json()
        assert claim2b["contact_id"] is not None
        assert claim2b["phone"] == "79997770002"


class TestOrphanedCallRecovery:
    """Если процесс агента умирает посреди звонка (например, его убили ради
    рестарта на новую версию кода), контакт навсегда остаётся в 'calling',
    а его кампания — в dispatch_state='running', и раньше НИКОГДА больше
    не подхватывалась (poll искал только 'requested'). poll() теперь сам
    восстанавливает такие зависшие звонки."""

    def _running_campaign_with_calling_contact(self, client, portal_token, phone):
        content = _xlsx_bytes(["Имя", "Телефон"], [["Кандидат", phone]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Осиротевший звонок", "scenario_id": SCENARIO, "source": "manual"},
        )
        cid = r.json()["campaign_id"]
        auth = {"X-Auth-Token": portal_token}
        client.post(f"/voicecall/campaigns/{cid}/start-dispatch", headers=auth)
        client.get("/voicecall/dispatch/poll", headers=auth)  # requested -> running
        claim = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                             headers=auth).json()
        return cid, claim["contact_id"], auth

    def test_fresh_calling_contact_not_touched(self, client, portal_token, main_module):
        """Звонок только что начался (last_attempt_at свежий) — poll не должен
        трогать его, кампания без ДРУГИХ pending-контактов не переподхватывается."""
        cid, contact_id, auth = self._running_campaign_with_calling_contact(
            client, portal_token, "79995551001")

        r = client.get("/voicecall/dispatch/poll", headers=auth)
        assert r.json()["campaign_id"] is None

        contacts = client.get("/voicecall/contacts", params={"campaign_id": cid}).json()["items"]
        assert contacts[0]["status"] == "calling"

    def test_stale_calling_contact_recovered_and_campaign_resumed(self, client, portal_token, main_module):
        """last_attempt_at старше 10 минут — считаем что агент умер посреди
        звонка, возвращаем контакт в очередь и переподхватываем кампанию."""
        cid, contact_id, auth = self._running_campaign_with_calling_contact(
            client, portal_token, "79995551002")

        old_ts = "2000-01-01T00:00:00"
        with main_module.db() as conn:
            conn.execute("UPDATE voicecall_contacts SET last_attempt_at=? WHERE id=?",
                         (old_ts, contact_id))

        r = client.get("/voicecall/dispatch/poll", headers=auth)
        assert r.status_code == 200
        assert r.json()["campaign_id"] == cid

        contacts = client.get("/voicecall/contacts", params={"campaign_id": cid}).json()["items"]
        assert contacts[0]["status"] == "pending"

        # И его можно снова забрать на звонок как обычно
        claim = client.post("/voicecall/dispatch/claim", params={"campaign_id": cid},
                             headers=auth).json()
        assert claim["contact_id"] == contact_id


class TestFunnelAndExport:
    def test_funnel_counts(self, client):
        headers = ["Имя", "Телефон", "Пол"]
        content = _xlsx_bytes(headers, [
            ["А", "79991110000", "женский"],   # screened_out
            ["Б", "79991110001", ""],           # pending
        ])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Воронка тест", "scenario_id": SCENARIO, "source": "manual"},
        )
        cid = r.json()["campaign_id"]
        funnel = client.get(f"/voicecall/campaigns/{cid}/funnel").json()
        assert funnel["loaded"] == 2
        assert funnel["screened_out"] == 1
        assert funnel["pending"] == 1

    def test_export_returns_valid_xlsx(self, client):
        content = _xlsx_bytes(["Имя", "Телефон"], [["В", "79991110002"]])
        r = client.post(
            "/voicecall/upload-contacts",
            files={"file": ("t.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": "Экспорт тест", "scenario_id": SCENARIO, "source": "manual"},
        )
        cid = r.json()["campaign_id"]
        r = client.get(f"/voicecall/campaigns/{cid}/export")
        assert r.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        assert headers[:2] == ["Имя", "Телефон"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert rows[0][0] == "В"
