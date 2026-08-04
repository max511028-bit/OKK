"""HTTP-слой Тендер-радара. Подключается в main.py одной строкой:

    from tenders_api import router as tenders_router
    app.include_router(tenders_router)

Отдельным модулем, а не внутри main.py, сознательно: main.py уже около
шести тысяч строк, и класть туда ещё один продукт целиком — верный способ
сделать его нечитаемым.

Все эндпоинты закрыты паролем портала (тем же, что конструктор обзвона).
"""
from __future__ import annotations

import datetime as _dt
import io
import threading
from typing import Any, Optional

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import Response
from pydantic import BaseModel

import tenders_pipeline as pipeline
from tenders_db import (
    JSON_FIELDS_DIRECTION,
    db,
    dumps,
    get_setting,
    init_db,
    now_iso,
    row_to_direction,
    set_setting,
)

router = APIRouter(prefix="/tenders", tags=["tenders"])

_READY = {"done": False}
_READY_LOCK = threading.Lock()


def _ensure_ready() -> None:
    """Ленивая инициализация: схема + реестр площадок. Дешевле, чем
    городить startup-хук в общем приложении портала."""
    if _READY["done"]:
        return
    with _READY_LOCK:
        if _READY["done"]:
            return
        init_db()
        try:
            pipeline.sync_sources()
        except Exception:  # noqa: BLE001 — без реестра API всё равно работает
            pass
        _READY["done"] = True


def _warmup() -> None:
    """Прогреть схему и реестр в фоне при старте портала.

    04.08, боевой деплой: первые ДВА запроса к вкладке вернули 502, а
    третий и дальше — 200 за 0.2с. Причина — ленивая инициализация
    (создание схемы + синхронизация 12 площадок) пришлась на первый
    запрос, и на одном ядре VPS она не уложилась в таймаут nginx.
    Пользователь, открывший вкладку первым, видел бы ошибку.

    Демон-поток, а не startup-хук: падение прогрева не должно мешать
    порталу стартовать."""
    def _run():
        try:
            _ensure_ready()
        except Exception:  # noqa: BLE001
            pass
    threading.Thread(target=_run, daemon=True, name="tenders-warmup").start()


_warmup()


def _check_password(request: Request) -> None:
    """Тот же пароль портала, что у конструктора обзвона."""
    from main import _vcs_check_password  # локальный импорт: избегаем цикла
    _vcs_check_password(request)


def _guard(request: Request) -> None:
    _check_password(request)
    _ensure_ready()


# ==========================================================================
# Справочники: группы, направления, ключевые слова
# ==========================================================================
class GroupIn(BaseModel):
    name: str
    description: str = ""
    color: str = "#4b7bec"
    sort_order: int = 0
    is_active: bool = True


class DirectionIn(BaseModel):
    group_id: Optional[int] = None
    name: str
    description: str = ""
    is_active: bool = True
    sort_order: int = 0
    min_score: float = 1.0
    min_price: Optional[float] = None
    max_price: Optional[float] = None
    regions: list[str] = []
    cities: list[str] = []
    customers: list[str] = []
    laws: list[str] = []
    okpd2: list[str] = []
    source_codes: list[str] = []


class KeywordIn(BaseModel):
    phrase: str
    kind: str = "include"       # include | exclude | require
    weight: float = 1.0
    match_mode: str = "stem"    # stem | exact | regex
    is_active: bool = True


@router.get("/groups")
def list_groups(request: Request):
    """Дерево «группа → направления» — то, ради чего всё затевалось."""
    _guard(request)
    with db() as conn:
        groups = [dict(r) for r in conn.execute(
            "SELECT * FROM groups ORDER BY sort_order, id").fetchall()]
        dirs = [row_to_direction(r) for r in conn.execute(
            "SELECT * FROM directions ORDER BY sort_order, id").fetchall()]
        counts = {r["direction_id"]: r["n"] for r in conn.execute(
            "SELECT direction_id, COUNT(*) AS n FROM keywords GROUP BY direction_id")}
        found = {r["direction_id"]: r["n"] for r in conn.execute(
            "SELECT direction_id, COUNT(*) AS n FROM matches GROUP BY direction_id")}
    for d in dirs:
        d["keywords_count"] = counts.get(d["id"], 0)
        d["tenders_count"] = found.get(d["id"], 0)
    by_group: dict[Any, list] = {}
    for d in dirs:
        by_group.setdefault(d["group_id"], []).append(d)
    for g in groups:
        g["is_active"] = bool(g["is_active"])
        g["directions"] = by_group.pop(g["id"], [])
    orphans = [d for lst in by_group.values() for d in lst]
    return {"groups": groups, "orphans": orphans}


@router.post("/groups")
def create_group(req: GroupIn, request: Request):
    _guard(request)
    with db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO groups(name, description, color, sort_order, is_active, created_at) "
                "VALUES(?,?,?,?,?,?)",
                (req.name.strip(), req.description, req.color, req.sort_order,
                 int(req.is_active), now_iso()))
        except Exception as e:
            raise HTTPException(400, f"Не удалось создать группу: {e}")
        return {"ok": True, "id": cur.lastrowid}


@router.put("/groups/{gid}")
def update_group(gid: int, req: GroupIn, request: Request):
    _guard(request)
    with db() as conn:
        conn.execute(
            "UPDATE groups SET name=?, description=?, color=?, sort_order=?, is_active=? "
            "WHERE id=?",
            (req.name.strip(), req.description, req.color, req.sort_order,
             int(req.is_active), gid))
    return {"ok": True}


@router.delete("/groups/{gid}")
def delete_group(gid: int, request: Request):
    """Группа удаляется, направления внутри остаются — уезжают в «Без
    группы». Терять настроенные правила из-за удаления папки нельзя."""
    _guard(request)
    with db() as conn:
        conn.execute("UPDATE directions SET group_id=NULL WHERE group_id=?", (gid,))
        conn.execute("DELETE FROM groups WHERE id=?", (gid,))
    return {"ok": True}


@router.get("/directions/{did}")
def get_direction(did: int, request: Request):
    _guard(request)
    with db() as conn:
        row = conn.execute("SELECT * FROM directions WHERE id=?", (did,)).fetchone()
        if not row:
            raise HTTPException(404, "Направление не найдено")
        d = row_to_direction(row)
        d["keywords"] = [dict(k) for k in conn.execute(
            "SELECT * FROM keywords WHERE direction_id=? ORDER BY id", (did,)).fetchall()]
        for k in d["keywords"]:
            k["is_active"] = bool(k["is_active"])
    return d


def _direction_values(req: DirectionIn) -> list:
    return [req.group_id, req.name.strip(), req.description, int(req.is_active),
            req.sort_order, req.min_score, req.min_price, req.max_price,
            dumps(req.regions), dumps(req.cities), dumps(req.customers),
            dumps(req.laws), dumps(req.okpd2), dumps(req.source_codes)]


@router.post("/directions")
def create_direction(req: DirectionIn, request: Request):
    _guard(request)
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO directions(group_id, name, description, is_active, sort_order, "
            "min_score, min_price, max_price, regions, cities, customers, laws, okpd2, "
            "source_codes, created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (*_direction_values(req), now_iso()))
        return {"ok": True, "id": cur.lastrowid}


@router.put("/directions/{did}")
def update_direction(did: int, req: DirectionIn, request: Request):
    _guard(request)
    with db() as conn:
        conn.execute(
            "UPDATE directions SET group_id=?, name=?, description=?, is_active=?, "
            "sort_order=?, min_score=?, min_price=?, max_price=?, regions=?, cities=?, "
            "customers=?, laws=?, okpd2=?, source_codes=? WHERE id=?",
            (*_direction_values(req), did))
    return {"ok": True}


@router.delete("/directions/{did}")
def delete_direction(did: int, request: Request):
    _guard(request)
    with db() as conn:
        conn.execute("DELETE FROM directions WHERE id=?", (did,))
    return {"ok": True}


@router.put("/directions/{did}/keywords")
def replace_keywords(did: int, items: list[KeywordIn], request: Request):
    """Список слов сохраняется целиком — так проще и в UI, и здесь:
    редактор отдаёт итоговое состояние, а не поштучные правки."""
    _guard(request)
    with db() as conn:
        if not conn.execute("SELECT 1 FROM directions WHERE id=?", (did,)).fetchone():
            raise HTTPException(404, "Направление не найдено")
        conn.execute("DELETE FROM keywords WHERE direction_id=?", (did,))
        for k in items:
            phrase = (k.phrase or "").strip()
            if not phrase:
                continue
            conn.execute(
                "INSERT INTO keywords(direction_id, phrase, kind, weight, match_mode, is_active) "
                "VALUES(?,?,?,?,?,?)",
                (did, phrase, k.kind, k.weight, k.match_mode, int(k.is_active)))
    return {"ok": True}


# ==========================================================================
# Площадки
# ==========================================================================
class SourceIn(BaseModel):
    is_enabled: Optional[bool] = None
    login: Optional[str] = None
    password: Optional[str] = None
    notes: Optional[str] = None


@router.get("/sources")
def list_sources(request: Request):
    _guard(request)
    with db() as conn:
        rows = [dict(r) for r in conn.execute("SELECT * FROM sources ORDER BY id").fetchall()]
    for r in rows:
        r["is_enabled"] = bool(r["is_enabled"])
        r["requires_auth"] = bool(r["requires_auth"])
        r.pop("password", None)      # наружу пароль площадки не отдаём
        r["running"] = pipeline.is_running(r["code"])
    return {"items": rows}


@router.put("/sources/{code}")
def update_source(code: str, req: SourceIn, request: Request):
    _guard(request)
    sets, params = [], []
    if req.is_enabled is not None:
        sets.append("is_enabled=?"); params.append(int(req.is_enabled))
    if req.login is not None:
        sets.append("login=?"); params.append(req.login)
    if req.password is not None:
        sets.append("password=?"); params.append(req.password)
    if req.notes is not None:
        sets.append("notes=?"); params.append(req.notes)
    if not sets:
        return {"ok": True}
    with db() as conn:
        conn.execute(f"UPDATE sources SET {', '.join(sets)} WHERE code=?", (*params, code))
    return {"ok": True}


# ==========================================================================
# Поиск: расписание и ручной запуск
# ==========================================================================
class ScanIn(BaseModel):
    """Ручной прогон. Всё необязательное: пусто — значит «все площадки,
    все направления, обычная глубина»."""
    source_code: Optional[str] = None
    direction_id: Optional[int] = None
    days: Optional[int] = None
    location: Optional[str] = None   # vps | home — чей сборщик отрабатывает


class SettingsIn(BaseModel):
    scan_interval_minutes: Optional[int] = None
    scan_enabled: Optional[bool] = None


@router.get("/settings")
def read_settings(request: Request):
    _guard(request)
    with db() as conn:
        return {
            "scan_interval_minutes": int(get_setting(conn, "scan_interval_minutes", "180")),
            "scan_enabled": get_setting(conn, "scan_enabled", "1") == "1",
            "last_scan_at": get_setting(conn, "last_scan_at", ""),
            "next_scan_at": get_setting(conn, "next_scan_at", ""),
            "home_collector_seen_at": get_setting(conn, "home_collector_seen_at", ""),
        }


@router.put("/settings")
def write_settings(req: SettingsIn, request: Request):
    _guard(request)
    with db() as conn:
        if req.scan_interval_minutes is not None:
            minutes = max(5, min(int(req.scan_interval_minutes), 7 * 24 * 60))
            set_setting(conn, "scan_interval_minutes", minutes)
        if req.scan_enabled is not None:
            set_setting(conn, "scan_enabled", "1" if req.scan_enabled else "0")
    return read_settings(request)


@router.post("/scan")
def scan_now(req: ScanIn, request: Request):
    """Запустить поиск руками. Работает в фоне: площадки отвечают
    медленно, и держать ради этого HTTP-запрос незачем — результат
    смотрится в журнале прогонов."""
    _guard(request)
    days = req.days if req.days is None else max(1, min(int(req.days), 365))

    def _work():
        try:
            if req.source_code:
                pipeline.run_source(req.source_code, triggered_by="manual",
                                    direction_id=req.direction_id, depth_days=days)
            else:
                pipeline.run_all(triggered_by="manual", direction_id=req.direction_id,
                                 depth_days=days, location=req.location)
        except Exception:  # noqa: BLE001
            pass

    threading.Thread(target=_work, daemon=True, name="tenders-manual-scan").start()
    return {"ok": True, "started": True, "days": days,
            "direction_id": req.direction_id, "source_code": req.source_code}


@router.post("/rematch")
def rematch(request: Request):
    """Пересчитать совпадения после правки направлений."""
    _guard(request)
    return pipeline.rematch_all()


@router.get("/runs")
def list_runs(request: Request, limit: int = 60):
    _guard(request)
    with db() as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM runs ORDER BY started_at DESC, id DESC LIMIT ?",
            (max(1, min(limit, 500)),)).fetchall()]
    return {"items": rows}


# ==========================================================================
# Найденные тендеры
# ==========================================================================
@router.get("/list")
def list_tenders(request: Request, direction_id: Optional[int] = None,
                 group_id: Optional[int] = None, status: Optional[str] = None,
                 source_code: Optional[str] = None, q: Optional[str] = None,
                 days: Optional[int] = None, limit: int = 200, offset: int = 0):
    _guard(request)
    # Показываем только то, что подходит хотя бы под одно направление.
    # Тендеры, переставшие подходить после правки фильтров, остаются в
    # базе без совпадений — так правка обратима (см. pipeline.rematch_all).
    where, params = ["EXISTS(SELECT 1 FROM matches mm WHERE mm.tender_id=t.id)"], []
    if status:
        where.append("t.status=?"); params.append(status)
    if source_code:
        where.append("t.source_code=?"); params.append(source_code)
    if q:
        where.append("(t.title LIKE ? OR t.customer LIKE ? OR t.description LIKE ?)")
        params += [f"%{q}%"] * 3
    if days:
        since = (_dt.datetime.now() - _dt.timedelta(days=int(days))).isoformat(timespec="seconds")
        where.append("COALESCE(t.published_at, t.first_seen_at) >= ?"); params.append(since)
    if direction_id:
        where.append("EXISTS(SELECT 1 FROM matches m WHERE m.tender_id=t.id AND m.direction_id=?)")
        params.append(direction_id)
    if group_id:
        where.append("EXISTS(SELECT 1 FROM matches m JOIN directions d ON d.id=m.direction_id "
                     "WHERE m.tender_id=t.id AND d.group_id=?)")
        params.append(group_id)
    sql_where = " AND ".join(where)
    with db() as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM tenders t WHERE {sql_where}",
                             params).fetchone()[0]
        rows = conn.execute(
            f"SELECT t.* FROM tenders t WHERE {sql_where} "
            f"ORDER BY COALESCE(t.published_at, t.first_seen_at) DESC, t.id DESC "
            f"LIMIT ? OFFSET ?", (*params, max(1, min(limit, 1000)), max(0, offset))).fetchall()
        ids = [r["id"] for r in rows]
        marks: dict[int, list] = {}
        if ids:
            qmarks = ",".join("?" * len(ids))
            for m in conn.execute(
                    f"SELECT m.tender_id, m.score, d.id AS did, d.name, g.name AS group_name "
                    f"FROM matches m JOIN directions d ON d.id=m.direction_id "
                    f"LEFT JOIN groups g ON g.id=d.group_id "
                    f"WHERE m.tender_id IN ({qmarks}) ORDER BY m.score DESC", ids):
                marks.setdefault(m["tender_id"], []).append(
                    {"direction_id": m["did"], "name": m["name"],
                     "group_name": m["group_name"], "score": m["score"]})
    items = []
    for r in rows:
        d = dict(r)
        d.pop("raw", None)
        d.pop("description", None)      # в списке не нужно, экономим трафик
        d["directions"] = marks.get(r["id"], [])
        items.append(d)
    return {"items": items, "total": total, "limit": limit, "offset": offset}


class TenderPatch(BaseModel):
    status: Optional[str] = None
    note: Optional[str] = None


@router.get("/item/{tid}")
def get_tender(tid: int, request: Request):
    _guard(request)
    with db() as conn:
        row = conn.execute("SELECT * FROM tenders WHERE id=?", (tid,)).fetchone()
        if not row:
            raise HTTPException(404, "Тендер не найден")
        d = dict(row)
        d["directions"] = [dict(m) for m in conn.execute(
            "SELECT m.score, m.matched, d.id AS direction_id, d.name, g.name AS group_name "
            "FROM matches m JOIN directions d ON d.id=m.direction_id "
            "LEFT JOIN groups g ON g.id=d.group_id WHERE m.tender_id=? ORDER BY m.score DESC",
            (tid,)).fetchall()]
    return d


@router.patch("/item/{tid}")
def patch_tender(tid: int, req: TenderPatch, request: Request):
    _guard(request)
    sets, params = [], []
    if req.status is not None:
        if req.status not in ("new", "interesting", "in_work", "rejected"):
            raise HTTPException(400, "Неизвестный статус")
        sets.append("status=?"); params.append(req.status)
    if req.note is not None:
        sets.append("note=?"); params.append(req.note)
    if not sets:
        return {"ok": True}
    sets.append("updated_at=?"); params.append(now_iso())
    with db() as conn:
        conn.execute(f"UPDATE tenders SET {', '.join(sets)} WHERE id=?", (*params, tid))
    return {"ok": True}


@router.get("/export")
def export_xlsx(request: Request, direction_id: Optional[int] = None,
                group_id: Optional[int] = None, status: Optional[str] = None,
                days: Optional[int] = None):
    """Выгрузка найденного в .xlsx — тем же openpyxl, что и отчёты обзвона."""
    _guard(request)
    data = list_tenders(request, direction_id=direction_id, group_id=group_id,
                        status=status, days=days, limit=1000)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Тендеры"
    headers = ["Опубликован", "Окончание", "Площадка", "Закон", "Заказчик", "Регион",
               "Название", "Цена", "Направления", "Балл", "Статус", "Ссылка"]
    ws.append(headers)
    for i, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4B7BEC")
        cell.alignment = Alignment(vertical="center")
    status_ru = {"new": "новый", "interesting": "интересен",
                 "in_work": "в работе", "rejected": "отклонён"}
    for t in data["items"]:
        ws.append([
            (t.get("published_at") or "")[:10], (t.get("deadline_at") or "")[:10],
            t.get("source_code", ""), t.get("law", ""), t.get("customer", ""),
            t.get("region", ""), t.get("title", ""), t.get("price"),
            ", ".join(d["name"] for d in t.get("directions", [])),
            t.get("best_score"), status_ru.get(t.get("status"), t.get("status")),
            t.get("url", ""),
        ])
    for col, width in zip("ABCDEFGHIJKL", (12, 12, 14, 8, 34, 22, 60, 14, 30, 8, 12, 40)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    stamp = _dt.datetime.now().strftime("%Y-%m-%d_%H-%M")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="tenders_{stamp}.xlsx"'})


@router.get("/summary")
def summary(request: Request):
    """Цифры для шапки вкладки."""
    _guard(request)
    with db() as conn:
        by_status = {r["status"]: r["n"] for r in conn.execute(
            "SELECT status, COUNT(*) AS n FROM tenders GROUP BY status")}
        total = sum(by_status.values())
        today = conn.execute(
            "SELECT COUNT(*) FROM tenders WHERE first_seen_at >= ?",
            (_dt.datetime.now().strftime("%Y-%m-%d"),)).fetchone()[0]
        sources_on = conn.execute(
            "SELECT COUNT(*) FROM sources WHERE is_enabled=1").fetchone()[0]
        dirs_on = conn.execute(
            "SELECT COUNT(*) FROM directions WHERE is_active=1").fetchone()[0]
        last_run = conn.execute(
            "SELECT started_at, source_code, status FROM runs "
            "ORDER BY started_at DESC LIMIT 1").fetchone()
    return {"total": total, "by_status": by_status, "today": today,
            "sources_enabled": sources_on, "directions_active": dirs_on,
            "last_run": dict(last_run) if last_run else None}
